"""
OpenPyXL - Excel

@author Aaric
@version 0.6.0-SNAPSHOT
"""

from openpyxl import load_workbook

wb = load_workbook("E:\\testdoc.xlsx")

sheets = wb.sheetnames

for sheet_name in sheets:
    sheet = wb[sheet_name]
    sheet["AC3"] = sheet.title
    print(sheet.title)

wb.save("E:\\testdoc2.xlsx")
